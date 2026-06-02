import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Raw data - loai bo dong rac (1, 2, 13)
raw = [
    "Build Institutional Grade Prediction Markets on Arc | Arc Blueprints",
    "How Arc Supports the Agentic Economy | Arc Blueprints",
    "How Arc Supports Lending and Borrowing | Arc Blueprints",
    "Introducing the ARC Whitepaper: Exploring Arc's Native Coordination Asset",
    "Agentic Economy on Arc",
    "Unified Balance Kit: One Integration for Unified USDC Flows",
    "How HIFI Offers Global Payouts with USDC, CPN, and CCTP",
    "App Kits: A Suite of SDKs to Build Onchain",
    "Open Sourcing Arc: Run Your Own Arc Node and Bug Bounty Program",
    "Running an Agentic Economic Flow on Arc with ERC-8183",
    "How Arc Supports Treasury Management | Arc Blueprints",
    "Arc's Quantum-Resistant Design and Roadmap: Why It Matters",
    "Preparing Blockchains for Q-Day",
    "Introducing Arc House and the Architects Program",
    "USDC on Arc: A Capital-Efficient Path for Banks",
    "Arc is proud to join the Mastercard Crypto Partner Program",
    "How Arc Supports Cross-Border Payments | Arc Blueprints",
    "Building the Internet Financial System: Circle's Product Vision for 2026",
    "Technical Insights on Arc Testnet Reliability",
    "Tokenizing Real-World Assets with Circle Contracts",
    "Beyond Stablecoins: The Rise of the Internet Financial System",
    "DoraHacks Start-up Ideas 2026: Pt.1 Digital Finance in the Circle/Arc ecosystem",
    "Introducing the Arc Builders Fund",
    "How to Build Real-Time Stablecoin FX in Your App with StableFX",
    "Circle Launches Arc Public Testnet",
]

src  = r"f:\Crape Linh Tinh\ArcHouse_Content.xlsx"
path = r"f:\Crape Linh Tinh\ArcHouse_Content2.xlsx"
import shutil; shutil.copy2(src, path)
wb = openpyxl.load_workbook(path)

if "External Content" in wb.sheetnames:
    del wb["External Content"]
ws = wb.create_sheet("External Content")

HDR_FILL = PatternFill("solid", fgColor="4B0082")
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
ODD_FILL  = PatternFill("solid", fgColor="F3E8FF")
EVEN_FILL = PatternFill("solid", fgColor="FAF5FF")

# Chi co 2 cot: # va Tieu de (khong co URL va Ngay vi external content khong lo duoc)
headers = ["#", "Tieu de"]
ws.append(headers)
ws.row_dimensions[1].height = 26
for col in range(1, 3):
    c = ws.cell(1, col)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

for i, tieude in enumerate(raw, 1):
    ws.append([i, tieude])
    fill = ODD_FILL if i % 2 == 1 else EVEN_FILL
    for col in range(1, 3):
        c = ws.cell(i + 1, col)
        c.fill = fill
        c.border = border
        c.font = Font(name="Calibri", size=11)
        c.alignment = Alignment(vertical="center", wrap_text=(col == 2))
    ws.cell(i+1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[i + 1].height = 20

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 75
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:B{len(raw)+1}"

wb.save(path)
import sys; sys.stdout.reconfigure(encoding='utf-8')
print(f"Done! Sheet 'External Content': {len(raw)} items")
