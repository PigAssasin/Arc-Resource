import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

videos = [
    ("Jun 1st, 2026", "Building Bermuda's Onchain Economy | The Government of Bermuda (David Burt) / Circle (Heath Tarbert, Kash Razzaghi)", "https://community.arc.io/home/videos/building-bermudas-onchain-economy-or-the-government-of-bermuda-david-burt-circle-heath-tarbert-kash-razzaghi-2026-06-01"),
    ("May 28th, 2026", "Agentic Demo Day: AI Projects Built on Arc", "https://community.arc.io/home/videos/agentic-demo-day-ai-projects-built-on-arc-2026-05-28"),
    ("May 28th, 2026", "Replay: Arc Builder Spotlight: TLAY - Machine-to-Machine Nanopayments on Arc", "https://community.arc.io/home/videos/replay-arc-builder-spotlight-tlay-machine-to-machine-nanopayments-on-arc-2026-05-28"),
    ("Apr 27th, 2026", "Scaling Institutional DeFi", "https://community.arc.io/home/videos/bridging-defi-and-tradfi-2026-04-27"),
    ("May 20th, 2026", "Replay: Arc Enterprise & DeFi Hackathon Spotlight: Chariot - Crosschain Collateral Lending Protocol on Arc", "https://community.arc.io/home/videos/replay-arc-enterprise-and-defi-hackathon-spotlight-chariot-crosschain-collateral-lending-protocol-on-arc-2026-05-20"),
    ("May 14th, 2026", "Circle Developer Grants: From idea to funded", "https://community.arc.io/home/videos/circle-developer-grants-from-idea-to-funded-2026-05-14"),
    ("Apr 27th, 2026", "Morpho x Arc featuring Merlin Egalite, Co-Founder of Morpho", "https://community.arc.io/home/videos/morpho-x-arc-2026-04-27"),
    ("May 1st, 2026", "Architect Technical Office Hours 4/30/26", "https://community.arc.io/home/videos/architect-technical-office-hours-2026-05-01"),
    ("May 1st, 2026", "ArcShop with Elton: Unified Balance Kit for Crosschain USDC Flows", "https://community.arc.io/home/videos/arcshop-with-elton-unified-balance-kit-for-crosschain-usdc-flows-2026-05-01"),
    ("Apr 29th, 2026", "ArcShop with HJ: Building Chain Agnostic Apps with Circle Gateway and Circle Wallets", "https://community.arc.io/home/videos/arcshop-with-hj-building-chain-agnostic-apps-with-circle-gateway-and-circle-wallets-2026-04-29"),
    ("Apr 21st, 2026", "The Arc developer experience: Hear from early builders at Sequence, Dynamic, and BuFi", "https://community.arc.io/home/videos/the-arc-developer-experience-hear-from-early-builders-at-sequence-dynamic-and-bufi-2026-04-21"),
    ("Apr 20th, 2026", "App Kits Developer Office Hours: Bridge, Swap, Send, and Monetization", "https://community.arc.io/home/videos/app-kits-developer-office-hours-bridge-swap-send-and-monetization-2026-04-20"),
    ("Apr 16th, 2026", "Why Circle Built Arc: The Vision and Key Features", "https://community.arc.io/home/videos/why-circle-built-arc-the-vision-and-key-features-2026-04-16"),
    ("Apr 3rd, 2026", "Event Replay: Introducing Arc House and Architects", "https://community.arc.io/home/videos/event-replay-introducing-arc-house-and-architects-2026-04-03"),
    ("Mar 26th, 2026", "Event Replay: Arc Community Spotlight: Social Payments with USDC Powered by XyloNet & PayX", "https://community.arc.io/home/videos/event-replay-arc-community-spotlight-social-payments-with-usdc-powered-by-xylonet-and-payx-2026-03-26"),
    ("Mar 19th, 2026", "Replay: USDC OpenClaw Hackathon Winner Spotlight: ClawRouter by BlockRunAI", "https://community.arc.io/home/videos/replay-usdc-openclaw-hackathon-winner-spotlight-clawrouter-by-blockrunai-2026-03-19"),
    ("Mar 13th, 2026", "Arc Day 1 Builder Series: Kosh", "https://community.arc.io/home/videos/arc-day-1-builder-series-kosh-2026-03-13"),
    ("Mar 14th, 2026", "Arc Studio: Buenos Aires", "https://community.arc.io/home/videos/arc-studio-buenos-aires-2026-03-14"),
    ("Mar 13th, 2026", "Arc Day One Builder Series: Peer", "https://community.arc.io/home/videos/arc-day-one-builder-series-peer-2026-03-13"),
    ("Feb 25th, 2026", "Arc Builders Fund Spotlight: Hibachi", "https://community.arc.io/home/videos/builders-fund-spotlight-hibachi-2026-02-25"),
    ("Feb 24th, 2026", "Day One: Blockradar", "https://community.arc.io/home/videos/day-one-blockradar-2026-02-24"),
    ("Feb 10th, 2026", "Espacio Cripto Podcast:", "https://community.arc.io/home/videos/espacio-cripto-podcast-2026-02-10"),
    ("Feb 5th, 2026", "Event Replay: Day One Architect: Para", "https://community.arc.io/home/videos/event-replay-day-one-architect-para-2026-02-05"),
    ("Feb 1st, 2026", "AI Agents, USDC, and the Programmable Economy | thirdweb (Furqan Rydhan) | Builder Series]", "https://community.arc.io/home/videos/ai-agents-usdc-and-the-programmable-economy-or-thirdweb-furqan-rydhan-or-builder-series-2026-02-01"),
    ("Jan 20th, 2026", "Emerging AI Trends with USDC", "https://community.arc.io/home/videos/emerging-ai-trends-with-usdc-2026-01-20"),
    ("Jan 20th, 2026", "Using Circle Developer Controlled Wallets to Send and Manage USDC", "https://community.arc.io/home/videos/using-circle-developer-controlled-wallets-to-send-and-manage-usdc-2026-01-20"),
    ("Jan 20th, 2026", "Using Circle Wallets to Send and Manage USDC", "https://community.arc.io/home/videos/using-circle-wallets-to-send-and-manage-usdc-2026-01-20"),
    ("Jan 20th, 2026", "How Circle Paymaster Works", "https://community.arc.io/home/videos/how-circle-paymaster-works-2026-01-20"),
    ("Jan 20th, 2026", "Introducing Circle Paymaster - Pay Gas Fees with USDC", "https://community.arc.io/home/videos/introducing-circle-paymaster-pay-gas-fees-with-usdc-2026-01-20"),
    ("Jan 20th, 2026", "CCTP vs. Gateway: What's the Difference and When to Use Each", "https://community.arc.io/home/videos/cctp-vs-gateway-whats-the-difference-and-when-to-use-each-2026-01-20"),
    ("Dec 21st, 2025", "Circle Ventures Spotlight: Trad.FI", "https://community.arc.io/home/videos/circle-ventures-spotlight-tradfi-2025-12-21"),
    ("Dec 19th, 2025", "Event Replay: Technical Insights on Arc Testnet Reliability", "https://community.arc.io/home/videos/event-replay-technical-insights-on-arc-testnet-reliability-2025-12-19"),
    ("Dec 18th, 2025", "How Does Circle Gateway Work? | Explained", "https://community.arc.io/home/videos/how-does-circle-gateway-work-or-explained-2025-12-18"),
    ("Dec 18th, 2025", "How Is USDC Unified Across Blockchains with Circle Gateway?", "https://community.arc.io/home/videos/how-is-usdc-unified-across-blockchains-with-circle-gateway-2025-12-18"),
    ("Dec 18th, 2025", "Demo: Moving USDC from Optimism to Ethereum with CCTP (Step-by-Step)", "https://community.arc.io/home/videos/demo-moving-usdc-from-optimism-to-ethereum-with-cctp-step-by-step-2025-12-18"),
    ("Dec 15th, 2025", "Event Replay: Crosschain Payments with CCTP", "https://community.arc.io/home/videos/event-replay-crosschain-payments-with-cctp-2025-12-15"),
    ("Dec 12th, 2025", "Event Replay: Day One Architect- Hinkal", "https://community.arc.io/home/videos/event-replay-day-one-architect-hinkal-2025-12-12"),
    ("Dec 11th, 2025", "Event Replay: Gateway with Blockradar", "https://community.arc.io/home/videos/event-replay-gateway-with-blockradar-2025-12-11"),
    ("Dec 10th, 2025", "Introducing CCTP Fast Transfer and How it Works", "https://community.arc.io/home/videos/introducing-cctp-fast-transfer-and-how-it-works-2025-12-10"),
    ("Dec 9th, 2025", "Event Replay: Arcshop- Introducing Bridge Kit", "https://community.arc.io/home/videos/arcshop-event-replay-introducing-bridge-kit-2025-12-09"),
    ("Dec 8th, 2025", "How Is USDC Interoperable? (Explained with Real-World Example)", "https://community.arc.io/home/videos/how-is-usdc-interoperable-explained-with-real-world-example-2025-12-08"),
    ("Dec 8th, 2025", "How to Use USDC in Real-World Payments Application (Part 2)", "https://community.arc.io/home/videos/how-to-use-usdc-in-real-world-payments-application-part-2-2025-12-08"),
    ("Dec 8th, 2025", "How to Use USDC in Real-World Payments Application (Part 1)", "https://community.arc.io/home/videos/how-to-use-usdc-in-real-world-payments-application-part-1-2025-12-08"),
    ("Dec 8th, 2025", "Using Circle Developer Controlled Wallets to Send and Manage USDC", "https://community.arc.io/home/videos/using-circle-developer-controlled-wallets-to-send-and-manage-usdc-2025-12-08"),
    ("Dec 8th, 2025", "Using Circle Wallets to Send and Manage USDC", "https://community.arc.io/home/videos/using-circle-wallets-to-send-and-manage-usdc-2025-12-08"),
    ("Dec 8th, 2025", "Arc Day One Spotlight: Instant Global Payments for Remote Workers with Hurupay", "https://community.arc.io/home/videos/arc-day-one-spotlight-instant-global-payments-for-remote-workers-with-hurupay-2025-12-08"),
    ("Dec 8th, 2025", "Arc Day One Spotlight: Simplifying Stablecoin Transactions with Blockradar", "https://community.arc.io/home/videos/arc-day-one-spotlight-simplifying-stablecoin-transactions-with-blockradar-2025-12-08"),
    ("Dec 8th, 2025", "Arc Day One Spotlight: Fast and Predictable Onchain Agentic Commerce with Crossmint", "https://community.arc.io/home/videos/arc-day-one-spotlight-fast-and-predictable-onchain-agentic-commerce-with-crossmint-2025-12-08"),
    ("Dec 6th, 2025", "Roundtable: Arc's Core Design Features", "https://community.arc.io/home/videos/roundtable-arcs-core-design-features-2025-12-06"),
    ("Dec 6th, 2025", "Roundtable: The Arc Experience", "https://community.arc.io/home/videos/roundtable-the-arc-experience-2025-12-06"),
    ("Dec 6th, 2025", "Roundtable: Arc the Economic OS w/Jeremy Allaire and Nikhil Chandhok Pt. 3", "https://community.arc.io/home/videos/roundtable-arc-the-economic-os-wjeremy-allaire-and-nikhil-chandhok-pt-3-2025-12-06"),
    ("Dec 6th, 2025", "Roundtable: Arc the Economic OS w/Jeremy Allaire and Nikhil Chandhok Pt. 2", "https://community.arc.io/home/videos/roundtable-arc-the-economic-os-wjeremy-allaire-and-nikhil-chandhok-pt-2-2025-12-06"),
    ("Dec 6th, 2025", "Roundtable: Arc the Economic OS w/Jeremy Allaire and Nikhil Chandhok Pt. 1", "https://community.arc.io/home/videos/roundtable-arc-the-economic-os-part-1-2025-12-06"),
    ("Dec 6th, 2025", "Event Replay: Welcome to Arc", "https://community.arc.io/home/videos/event-replay-welcome-to-arc"),
    ("Dec 6th, 2025", "Event Replay: Day One Architect: Crossmint", "https://community.arc.io/home/videos/event-replay-day-one-architect-crossmint"),
    ("Nov 13th, 2025", "Event Replay: Day One Architect Highlight: Blockradar", "https://community.arc.io/home/videos/event-replay-day-one-architect-highlight-blockradar-2025-11-13"),
    ("Nov 1st, 2025", "Making Your First USDC Transaction On Ethereum", "https://community.arc.io/home/videos/making-your-first-usdc-transaction-on-ethereum-2025-11-01"),
    ("Nov 1st, 2025", "How USDC Works On Ethereum: ERC-20 Token Standard", "https://community.arc.io/home/videos/how-usdc-works-on-ethereum-erc-20-token-standard-2025-11-01"),
    ("Nov 1st, 2025", "How USDC Works Onchain", "https://community.arc.io/home/videos/how-usdc-works-onchain-2025-11-01"),
    ("Nov 1st, 2025", "What is USDC?", "https://community.arc.io/home/videos/what-is-usdc-2025-11-01"),
    ("Nov 3rd, 2025", "Roundtable: Arc Global Connection", "https://community.arc.io/home/videos/roundtable-arc-global-connection-2025-11-03"),
    ("Nov 3rd, 2025", "Event Replay: Finality You Can Count On", "https://community.arc.io/home/videos/finality-you-can-count-on-2025-11-03"),
    ("Oct 30th, 2025", "Jeremy Allaire: Arc Public Testnet is live", "https://community.arc.io/home/videos/jeremy-allaire-arc-public-testnet-is-live-2025-10-30"),
    ("Oct 28th, 2025", "AI Agents on Arc with USDC", "https://community.arc.io/home/videos/ai-agents-on-arc-with-usdc-2025-10-28"),
    ("Oct 24th, 2025", "Arc: Rebuilding Money on the internet", "https://community.arc.io/home/videos/arc-rebuilding-money-on-the-internet-2025-10-24"),
    ("Oct 24th, 2025", "Why Circle Decided to Build Arc", "https://community.arc.io/home/videos/why-circle-decided-to-build-arc"),
    ("Oct 24th, 2025", "Arc: Developer Overview in 90 secs", "https://community.arc.io/home/videos/arc-developer-overview-in-90-secs-2025-10-24"),
    ("Oct 24th, 2025", "Are Stablecoins the Internet's Financial Infrastructure?", "https://community.arc.io/home/videos/are-stablecoins-the-internets-financial-infrastructure-2025-10-24"),
    ("Oct 24th, 2025", "What are Stablecoins?", "https://community.arc.io/home/videos/what-are-stablecoins-2025-10-24"),
    ("Oct 21st, 2025", "Arc: Developer Overview with Circle's Lead Product Manager Sanket Jain", "https://community.arc.io/home/videos/arc-developer-overview-with-circles-lead-product-manager-sanket-jain-2025-10-21"),
    ("Oct 10th, 2025", "Discover the Vision: Hear from Arc's Founding Team", "https://community.arc.io/home/videos/discover-the-vision-hear-from-arcs-founding-team-2025-10-10"),
]

# Load file Excel có sẵn
path = r"f:\Crape Linh Tinh\ArcHouse_LichSu.xlsx"
wb = openpyxl.load_workbook(path)

# Xóa sheet Video cũ nếu có
if "Video" in wb.sheetnames:
    del wb["Video"]

ws = wb.create_sheet("Video")

# Style
HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
VIDEO_FILL = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL  = PatternFill("solid", fgColor="EBF3FB")

# Header
headers = ["#", "Ngay", "Tieu de", "URL"]
ws.append(headers)
ws.row_dimensions[1].height = 26
for col in range(1, 5):
    c = ws.cell(1, col)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

# Data
for i, (ngay, tieude, url) in enumerate(videos, 1):
    ws.append([i, ngay, tieude, url])
    fill = VIDEO_FILL if i % 2 == 1 else ALT_FILL
    for col in range(1, 5):
        c = ws.cell(i + 1, col)
        c.fill = fill
        c.border = border
        c.font = Font(name="Calibri", size=11)
        c.alignment = Alignment(vertical="center", wrap_text=(col == 3))
    ws.cell(i + 1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(i + 1, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[i + 1].height = 20

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 17
ws.column_dimensions["C"].width = 70
ws.column_dimensions["D"].width = 55
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:D{len(videos)+1}"

out = r"f:\Crape Linh Tinh\ArcHouse_Content.xlsx"
wb.save(out)
import sys; sys.stdout.reconfigure(encoding='utf-8')
print(f"Done! Saved to: {out} ({len(videos)} videos)")
