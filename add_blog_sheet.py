import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

blogs = [
    ("May 30th, 2026", "Circle's Post-Quantum Security Roadmap: Securing blockchains, smart contracts, and digital assets for the quantum era", "https://community.arc.io/home/blogs/circles-post-quantum-security-roadmap-securing-blockchains-smart-contracts-and-digital-assets-for-the-quantum-era-2026-05-30"),
    ("May 26th, 2026", "Goldsky 🤝 Arc Builders Fund: real-time data infrastructure for onchain finance", "https://community.arc.io/home/blogs/goldsky-arc-builders-fund-real-time-data-infrastructure-for-onchain-finance-2026-05-26"),
    ("May 21st, 2026", "Stablecorp brings QCAD to Arc, expanding StableFX into Canadian dollars", "https://community.arc.io/home/blogs/stablecorp-brings-qcad-to-arc-expanding-stablefx-into-canadian-dollars-2026-05-21"),
    ("May 14th, 2026", "Circle Developer Grants Program Relaunches", "https://community.arc.io/home/blogs/circle-developer-grants-program-relaunches-2026-05-14"),
    ("Apr 2nd, 2026", "Tradable joins the Arc Builders Fund: institutional private credit, onchain", "https://community.arc.io/home/blogs/tradable-joins-the-arc-builders-fund-institutional-private-credit-onchain-2026-04-02"),
    ("Mar 27th, 2026", "TRM Labs joins Arc: blockchain intelligence + monitoring for enterprise-grade apps", "https://community.arc.io/home/blogs/trm-labs-joins-arc-blockchain-intelligence-monitoring-for-enterprise-grade-apps"),
    ("Mar 25th, 2026", "Across is live on Arc Testnet: day-one crosschain transfers for builders", "https://community.arc.io/home/blogs/arc-x-across"),
    ("Mar 14th, 2026", "Introducing Circle Skills: AI Tooling to Help Developers Integrate Faster", "https://community.arc.io/home/blogs/circle-ai-skills"),
    ("Feb 25th, 2026", "Arc x Elliptic: Blockchain analytics and monitoring for compliance-first Arc apps", "https://community.arc.io/home/blogs/arc-x-elliptic"),
    ("Feb 18th, 2026", "Arc x Alchemy: Alchemy Integrates with Arc", "https://community.arc.io/home/blogs/arc-x-alchemy"),
    ("Feb 13th, 2026", "Arc x Hibachi: Perpetuals on Arc, With ZK-Verified Settlement", "https://community.arc.io/home/blogs/arc-x-hibachi-spotlight"),
    ("Feb 4th, 2026", "Arc x QuickNode: Arc RPC infrastructure you can actually ship on", "https://community.arc.io/home/blogs/arc-x-quicknode"),
    ("Feb 3rd, 2026", "Quickstart Spotlight: Transfer USDC or EURC on Arc using Dev-Controlled Wallets", "https://community.arc.io/home/blogs/quickstart-spotlight-transfer-usdc-or-eurc-on-arc-using-dev-controlled-wallets"),
    ("Jan 29th, 2026", "Quickstart Spotlight: Bridge USDC to Arc with CCTP + Bridge Kit", "https://community.arc.io/home/blogs/quickstart-spotlight-bridge-usdc-to-arc-with-cctp-bridge-kit"),
    ("Jan 27th, 2026", "Quickstart Spotlight: Deploy an ERC-20 on Arc using Circle Contracts", "https://community.arc.io/home/blogs/quickstart-spotlight-deploy-an-erc-20-on-arc-using-circle-contracts"),
    ("Jan 21st, 2026", "Building an Autonomous Wallet Agent", "https://community.arc.io/home/blogs/building-an-autonomous-wallet-agent"),
    ("Nov 8th, 2025", "Welcome to the Arc Hub, A Community Introduction", "https://community.arc.io/home/blogs/welcome-to-the-arc-hub-an-introduction"),
    ("Nov 4th, 2025", "The Next Internet Economy, Built on Arc with USDC", "https://community.arc.io/home/blogs/the-next-internet-economy-built-on-arc-with-usdc-2025-11-04"),
    ("Oct 10th, 2025", "Arc's Deterministic Finality", "https://community.arc.io/home/blogs/httpswwwarcnetworkblogarcs-deterministic-finality-the-bespoke-consensus-layer-built-using-malachite"),
    ("Oct 10th, 2025", "How Gas Works on Arc", "https://community.arc.io/home/blogs/how-gas-works-on-arc-2025-10-10"),
    ("Oct 10th, 2025", "Deterministic Finality on Arc", "https://community.arc.io/home/blogs/deterministic-finality-on-arc"),
    ("Oct 10th, 2025", "Introducing Arc: The Economic OS for the internet", "https://community.arc.io/home/blogs/introducing-arc-an-open-layer-1-blockchain-purpose-built-for-stablecoin-finance"),
    ("Jan 15th, 2026", "Ship Stablecoin Apps Faster with App Kits", "https://community.arc.io/home/blogs/ship-stablecoin-apps-faster-app-kits"),
]

path = r"f:\Crape Linh Tinh\ArcHouse_Content.xlsx"
wb = openpyxl.load_workbook(path)

if "Blog" in wb.sheetnames:
    del wb["Blog"]
ws = wb.create_sheet("Blog")

HDR_FILL = PatternFill("solid", fgColor="1B4F2A")
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
EVEN_FILL = PatternFill("solid", fgColor="D5F5E3")
ODD_FILL  = PatternFill("solid", fgColor="EAFAF1")

headers = ["#", "Ngay", "Tieu de", "URL"]
ws.append(headers)
ws.row_dimensions[1].height = 26
for col in range(1, 5):
    c = ws.cell(1, col)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

for i, (ngay, tieude, url) in enumerate(blogs, 1):
    ws.append([i, ngay, tieude, url])
    fill = ODD_FILL if i % 2 == 1 else EVEN_FILL
    for col in range(1, 5):
        c = ws.cell(i + 1, col)
        c.fill = fill
        c.border = border
        c.font = Font(name="Calibri", size=11)
        c.alignment = Alignment(vertical="center", wrap_text=(col == 3))
    ws.cell(i+1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(i+1, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[i + 1].height = 20

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 17
ws.column_dimensions["C"].width = 70
ws.column_dimensions["D"].width = 55
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:D{len(blogs)+1}"

wb.save(path)
import sys; sys.stdout.reconfigure(encoding='utf-8')
print(f"Done! Sheet 'Blog' added: {len(blogs)} blogs")
