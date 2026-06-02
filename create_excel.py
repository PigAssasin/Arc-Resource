import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = [
    ("Watch a Video", "Jun 1st, 2026", "ArcShop with HJ: Building Chain Agnostic Apps with Circle Gateway and Circle Wallets"),
    ("Daily Active", "May 31st, 2026", ""),
    ("Watch a Video", "May 31st, 2026", "The Arc developer experience: Hear from early builders at Sequence, Dynamic, and BuFi"),
    ("Watch a Video", "May 31st, 2026", "App Kits Developer Office Hours: Bridge, Swap, Send, and Monetization"),
    ("Daily Active", "May 30th, 2026", ""),
    ("Watch a Video", "May 31st, 2026", "Arc Day 1 Builder Series: Kosh"),
    ("Watch a Video", "May 31st, 2026", "Arc Studio: Buenos Aires"),
    ("Read Content", "May 30th, 2026", "The Next Internet Economy, Built on Arc with USDC"),
    ("Read Content", "May 30th, 2026", "Welcome to the Arc Hub, A Community Introduction"),
    ("Read Content", "May 30th, 2026", "Arc 🤝 Turnkey: Wallet and signing infrastructure for builders on Arc"),
    ("Read Content", "May 30th, 2026", "Stablecorp brings QCAD to Arc, expanding StableFX into Canadian dollars"),
    ("Watch a Video", "May 30th, 2026", "Arc Builders Fund Spotlight: Hibachi"),
    ("Watch a Video", "May 30th, 2026", "Day One: Blockradar"),
    ("Daily Active", "May 29th, 2026", ""),
    ("Watch a Video", "May 30th, 2026", "CCTP vs. Gateway: What's the Difference and When to Use Each"),
    ("Watch a Video", "May 30th, 2026", "Circle Ventures Spotlight: Trad.FI"),
    ("Read Content", "May 29th, 2026", "Goldsky 🤝 Arc Builders Fund: real-time data infrastructure for onchain finance"),
    ("Read Content", "May 29th, 2026", "Ship Stablecoin Apps Faster with App Kits"),
    ("Read Content", "May 29th, 2026", "Arc Opens Its Code, Its Nodes, and a Formal Path to Break It Before Mainnet"),
    ("Read Content", "May 29th, 2026", "Tradable joins the Arc Builders Fund: institutional private credit, onchain"),
    ("Read Content", "May 29th, 2026", "Across is live on Arc Testnet: day-one crosschain transfers for builders"),
    ("Daily Active", "May 28th, 2026", ""),
    ("Watch a Video", "May 29th, 2026", "Event Replay: Arcshop- Introducing Bridge Kit"),
    ("Watch a Video", "May 29th, 2026", "How Is USDC Interoperable? (Explained with Real-World Example)"),
    ("Watch a Video", "May 28th, 2026", "How to Use USDC in Real-World Payments Application (Part 2)"),
    ("Watch a Video", "May 28th, 2026", "How to Use USDC in Real-World Payments Application (Part 1)"),
    ("Daily Active", "May 27th, 2026", ""),
    ("Watch a Video", "May 28th, 2026", "Arc Day One Spotlight: Simplifying Stablecoin Transactions with Blockradar"),
    ("Watch a Video", "May 28th, 2026", "Arc Day One Spotlight: Fast and Predictable Onchain Agentic Commerce with Crossmint"),
    ("Read Content", "May 27th, 2026", "TRM Labs joins Arc: blockchain intelligence + monitoring for enterprise-grade apps"),
    ("Read Content", "May 27th, 2026", "Arc x Elliptic: Blockchain analytics and monitoring for compliance-first Arc apps"),
    ("Watch a Video", "May 27th, 2026", "Roundtable: Arc's Core Design Features"),
    ("Watch a Video", "May 27th, 2026", "Roundtable: The Arc Experience"),
    ("Read Content", "May 27th, 2026", "Arc x Alchemy: Alchemy Integrates with Arc"),
    ("Daily Active", "May 26th, 2026", ""),
    ("Watch a Video", "May 27th, 2026", "Discover the Vision: Hear from Arc's Founding Team"),
    ("Read Content", "May 27th, 2026", "Arc's Deterministic Finality"),
    ("Read Content", "May 27th, 2026", "Arc 🤝 Dynamic: Better onboarding for apps built on Arc"),
    ("Watch a Video", "May 26th, 2026", "AI Agents on Arc with USDC"),
    ("Read Content", "May 26th, 2026", "How Gas Works on Arc"),
    ("Read Content", "May 26th, 2026", "Deterministic Finality on Arc"),
    ("Read Content", "May 26th, 2026", "Introducing Arc: The Economic OS for the internet"),
    ("Watch a Video", "May 26th, 2026", "Event Replay: Trustless USDC Agents on Arc"),
    ("Watch a Video", "May 26th, 2026", "Arc Enterprise & DeFi Hackathon Spotlight: Crumb - Gasless USDC Nanopayments and settlement on Arc"),
    ("Daily Active", "May 25th, 2026", ""),
    ("Read Content", "May 25th, 2026", "How Arc Supports the Agentic Economy | Arc Blueprints"),
    ("Read Content", "May 25th, 2026", "New Group: Introducing Agentic Economy in Arc House"),
    ("Watch a Video", "May 25th, 2026", "Building the Agentic Economy on Arc: VibeCard"),
    ("Watch a Video", "May 25th, 2026", "Day One: Crossmint"),
    ("Watch a Video", "May 25th, 2026", "Arc Enterprise & DeFi Hackathon Spotlight: Blink Nanopayment Insurance"),
    ("Daily Active", "May 24th, 2026", ""),
    ("Watch a Video", "May 24th, 2026", "Arc Day One Builder Series: Peer"),
    ("Daily Active", "May 23rd, 2026", ""),
    ("Read Content", "May 24th, 2026", "Arc's Quantum-Resistant Design and Roadmap: Why It Matters"),
    ("Read Content", "May 24th, 2026", "How Arc Supports Treasury Management | Arc Blueprints"),
    ("Read Content", "May 23rd, 2026", "Circle Launches Arc Public Testnet"),
    ("Read Content", "May 23rd, 2026", "DoraHacks Start-up Ideas 2026: Pt.1 Digital Finance in the Circle/Arc ecosystem"),
    ("Read Content", "May 23rd, 2026", "Quickstart Spotlight: Bridge USDC to Arc with CCTP + Bridge Kit"),
    ("Watch a Video", "May 23rd, 2026", "ArcShop with Elton: Unified Balance Kit for Crosschain USDC Flows"),
    ("Watch a Video", "May 23rd, 2026", "Introducing CCTP Fast Transfer and How it Works"),
    ("Watch a Video", "May 23rd, 2026", "Demo: Moving USDC from Optimism to Ethereum with CCTP (Step-by-Step)"),
    ("Daily Active", "May 22nd, 2026", ""),
    ("Watch a Video", "May 22nd, 2026", "Using Circle Wallets to Send and Manage USDC"),
    ("Watch a Video", "May 22nd, 2026", "Using Circle Developer Controlled Wallets to Send and Manage USDC"),
    ("Read Content", "May 22nd, 2026", "Building an Autonomous Wallet Agent"),
    ("Read Content", "May 22nd, 2026", "Guest Post: The New Era of Agentic Commerce, Highlights from the Arc Hackathon"),
    ("Watch a Video", "May 22nd, 2026", "AI Agents, USDC, and the Programmable Economy | thirdweb (Furqan Rydhan) | Builder Series]"),
    ("Read Content", "May 22nd, 2026", "Introducing Circle Skills: AI Tooling to Help Developers Integrate Faster"),
    ("Read Content", "May 22nd, 2026", "Circle Agent Stack - Builder Feedback Survey"),
    ("Daily Active", "May 21st, 2026", ""),
    ("Watch a Video", "May 22nd, 2026", "Emerging AI Trends with USDC"),
    ("Watch a Video", "May 21st, 2026", "Event Replay: Technical Insights on Arc Testnet Reliability"),
    ("Watch a Video", "May 21st, 2026", "Event Replay: Day One Architect: Para"),
    ("Watch a Video", "May 21st, 2026", "Event Replay: Arc Community Spotlight: Social Payments with USDC Powered by XyloNet & PayX"),
    ("Daily Active", "May 20th, 2026", ""),
    ("Watch a Video", "May 21st, 2026", "Morpho x Arc featuring Merlin Egalite, Co-Founder of Morpho"),
    ("Read Content", "May 21st, 2026", "Architects: Terms & Conditions"),
    ("Read Content", "May 21st, 2026", "Architects: Tiers & Benefits"),
    ("Read Content", "May 21st, 2026", "Arc Engagement Amplification Guide"),
    ("Watch a Video", "May 20th, 2026", "How USDC Works Onchain"),
    ("Watch a Video", "May 20th, 2026", "How Circle Paymaster Works"),
    ("Watch a Video", "May 20th, 2026", "Introducing Circle Paymaster - Pay Gas Fees with USDC"),
    ("Read Content", "May 20th, 2026", "We Made Circle Docs AI-Native"),
    ("Read Content", "May 20th, 2026", "What Nanopayments powered by Circle Gateway changes for Arc builders"),
    ("Daily Active", "May 19th, 2026", ""),
    ("Watch a Video", "May 20th, 2026", "Event Replay: Introducing Arc House and Architects"),
    ("Read Content", "May 20th, 2026", "Architects: Roles"),
    ("Daily Active", "May 19th, 2026", ""),
    ("Watch a Video", "May 19th, 2026", "Event Replay: Building an Agentic Economy on Arc with RSoft Agentic Bank"),
    ("Watch a Video", "May 19th, 2026", "Running an Agentic Economic Flow on Arc with ERC-8183"),
    ("Watch a Video", "May 19th, 2026", "Circle Agent Stack Quickstart - financial infrastructure for the agentic economy."),
    ("Watch a Video", "May 19th, 2026", "Circle Developer Grants: From idea to funded"),
    ("Read Content", "May 19th, 2026", "Circle Developer Grants Program Relaunches"),
    ("Read Content", "May 19th, 2026", "How Arc Supports Lending and Borrowing | Arc Blueprints"),
    ("Read Content", "May 19th, 2026", "Build Institutional Grade Prediction Markets on Arc | Arc Blueprints"),
    ("Read Content", "May 19th, 2026", "Arc 🤝 LI.FI: Crosschain routing and liquidity access for Arc builders"),
    ("Read Content", "May 19th, 2026", "Introducing the ARC Whitepaper: Exploring Arc's Native Coordination Asset"),
]

# Màu sắc
COLOR_HEADER = "1E3A5F"       # xanh đậm
COLOR_VIDEO  = "D6E4F0"       # xanh nhạt
COLOR_READ   = "D5F5E3"       # xanh lá nhạt
COLOR_DAILY  = "FEF9E7"       # vàng nhạt
COLOR_ROW_ALT= "F8F9FA"       # xám rất nhạt (row chẵn)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Lịch sử Arc House"

# Header
headers = ["#", "Loại hoạt động", "Ngày", "Tiêu đề"]
ws.append(headers)

header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
header_fill  = PatternFill("solid", fgColor=COLOR_HEADER)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.font   = header_font
    cell.fill   = header_fill
    cell.alignment = header_align

ws.row_dimensions[1].height = 28

# Data rows
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, (loai, ngay, tieu_de) in enumerate(data, 1):
    row_num = i + 1
    ws.append([i, loai, ngay, tieu_de])

    # Màu theo loại
    if loai == "Watch a Video":
        fill_color = COLOR_VIDEO
    elif loai == "Read Content":
        fill_color = COLOR_READ
    elif loai == "Daily Active":
        fill_color = COLOR_DAILY
    else:
        fill_color = COLOR_ROW_ALT if i % 2 == 0 else "FFFFFF"

    row_fill = PatternFill("solid", fgColor=fill_color)

    for col in range(1, 5):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = row_fill
        cell.border = border
        cell.font   = Font(name="Calibri", size=11)
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))

    # Căn giữa cột số thứ tự và loại
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row_num].height = 20

# Độ rộng cột
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 75

# Freeze header
ws.freeze_panes = "A2"

# Auto filter
ws.auto_filter.ref = f"A1:D{len(data)+1}"

# Sheet thống kê
ws2 = wb.create_sheet("Thống kê")
ws2.append(["Loại", "Số lượng"])
from collections import Counter
counts = Counter(row[0] for row in data)
for loai, count in sorted(counts.items()):
    ws2.append([loai, count])
ws2.append(["TỔNG", len(data)])

# Format sheet thống kê
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.font = Font(name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="center")
ws2.cell(1,1).font = Font(bold=True, size=12)
ws2.cell(1,2).font = Font(bold=True, size=12)
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 12

out = r"f:\Crape Linh Tinh\ArcHouse_LichSu.xlsx"
wb.save(out)
import sys
sys.stdout.reconfigure(encoding='utf-8')
print(f"Done: {out}")
print(f"Total: {len(data)} rows")
for loai, count in sorted(counts.items()):
    print(f"  {loai}: {count}")
