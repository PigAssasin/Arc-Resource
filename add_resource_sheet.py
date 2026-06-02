import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

resources = [
    ("May 26th, 2026", "Developer Brand Survey", "https://community.arc.io/home/resources/circle-developer-brand-survey-2026-05-26"),
    ("May 26th, 2026", "Architects: Program Overview", "https://community.arc.io/home/resources/architects-overview"),
    ("May 26th, 2026", "Architects: Tiers & Benefits", "https://community.arc.io/home/resources/architects-tiers-and-benefits"),
    ("May 26th, 2026", "Architects: Roles", "https://community.arc.io/home/resources/architects-roles"),
    ("May 26th, 2026", "Arc Engagement Amplification Guide", "https://community.arc.io/home/resources/arc-engagement-amplification-guide"),
    ("May 26th, 2026", "Architects: Contribution Opportunities", "https://community.arc.io/home/resources/architects-contribution-opportunities"),
    ("May 26th, 2026", "Architects: Terms & Conditions", "https://community.arc.io/home/resources/architects-terms-and-conditions"),
]

path = r"f:\Crape Linh Tinh\ArcHouse_Content.xlsx"
wb = openpyxl.load_workbook(path)

if "Resource" in wb.sheetnames:
    del wb["Resource"]
ws = wb.create_sheet("Resource")

HDR_FILL = PatternFill("solid", fgColor="7C4D0C")
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
ODD_FILL  = PatternFill("solid", fgColor="FEF3E2")
EVEN_FILL = PatternFill("solid", fgColor="FDF8F0")

headers = ["#", "Ngay", "Tieu de", "URL"]
ws.append(headers)
ws.row_dimensions[1].height = 26
for col in range(1, 5):
    c = ws.cell(1, col)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

for i, (ngay, tieude, url) in enumerate(resources, 1):
    ws.append([i, ngay, tieude, url])
    fill = ODD_FILL if i % 2 == 1 else EVEN_FILL
    for col in range(1, 5):
        c = ws.cell(i + 1, col)
        c.fill = fill
        c.border = border
        c.font = Font(name="Calibri", size=11)
        c.alignment = Alignment(vertical="center", wrap_text=(col == 3))
    ws.cell(i+1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(i+1, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[i + 1].height = 20

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 17
ws.column_dimensions["C"].width = 55
ws.column_dimensions["D"].width = 60
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:D{len(resources)+1}"

wb.save(path)
import sys; sys.stdout.reconfigure(encoding='utf-8')
print(f"Done! Sheet 'Resource' added: {len(resources)} items")
